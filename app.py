# ============================================================
#  QA Buddy — app.py
#  Modern PDF generation with ReportLab Platypus
# ============================================================

from flask import Flask, render_template, request, redirect, url_for, send_file
import sqlite3
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile
import shutil
import os
import json
from datetime import datetime

# ReportLab — Platypus (flowable layout engine)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether
)

import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from playwright.sync_api import sync_playwright


# ── Colour palette ────────────────────────────────────────────
C_DARK   = colors.HexColor('#18181b')
C_MID    = colors.HexColor('#3f3f46')
C_ACCENT = colors.HexColor('#2563eb')
C_AMBER  = colors.HexColor('#d97706')
C_GREEN  = colors.HexColor('#16a34a')
C_RED    = colors.HexColor('#dc2626')
C_BORDER = colors.HexColor('#e4e4e7')
C_BG     = colors.HexColor('#fafafa')
C_WHITE  = colors.white
C_SUBTLE = colors.HexColor('#71717a')


# ── Shared PDF helpers ────────────────────────────────────────

def build_styles():
    return {
        'title': ParagraphStyle('DocTitle',
            fontName='Helvetica-Bold', fontSize=22,
            textColor=C_WHITE, alignment=TA_LEFT, spaceAfter=2),
        'sub': ParagraphStyle('DocSub',
            fontName='Helvetica', fontSize=10,
            textColor=colors.HexColor('#d4d4d8'), alignment=TA_LEFT),
        'heading': ParagraphStyle('SectionH',
            fontName='Helvetica-Bold', fontSize=11,
            textColor=C_DARK, spaceBefore=14, spaceAfter=5),
        'body': ParagraphStyle('Body',
            fontName='Helvetica', fontSize=10,
            textColor=C_MID, leading=16, spaceAfter=4),
        'label': ParagraphStyle('Label',
            fontName='Helvetica-Bold', fontSize=8,
            textColor=C_SUBTLE, spaceAfter=3, spaceBefore=6),
        'th': ParagraphStyle('TH',
            fontName='Helvetica-Bold', fontSize=9,
            textColor=C_WHITE, alignment=TA_CENTER),
        'td': ParagraphStyle('TD',
            fontName='Helvetica', fontSize=9,
            textColor=C_MID, alignment=TA_CENTER, leading=13),
    }


def pdf_header(title, subtitle, accent=None):
    """Coloured header banner."""
    accent = accent or C_DARK
    S = build_styles()
    tbl = Table(
        [[Paragraph(title, S['title'])],
         [Paragraph(subtitle, S['sub'])]],
        colWidths=[170*mm]
    )
    tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,-1), accent),
        ('TOPPADDING',   (0,0), (-1,-1), 18),
        ('BOTTOMPADDING',(0,0), (-1,-1), 18),
        ('LEFTPADDING',  (0,0), (-1,-1), 20),
        ('RIGHTPADDING', (0,0), (-1,-1), 20),
    ]))
    return tbl


def section_card(label, content, accent=None):
    """A labelled content card with coloured left border."""
    accent = accent or C_ACCENT
    S = build_styles()
    rows = [[Paragraph(label.upper(), S['label'])]]
    for line in content.strip().split('\n'):
        if line.strip():
            rows.append([Paragraph(line, S['body'])])
    tbl = Table(rows, colWidths=[165*mm])
    tbl.setStyle(TableStyle([
        ('LEFTPADDING',   (0,0), (-1,-1), 12),
        ('RIGHTPADDING',  (0,0), (-1,-1), 10),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LINEBEFORE',    (0,0), (0,-1),  3, accent),
        ('BACKGROUND',    (0,0), (-1,-1), C_BG),
        ('BOX',           (0,0), (-1,-1), 0.5, C_BORDER),
    ]))
    return KeepTogether([tbl, Spacer(1, 10)])


def data_table(headers, rows, col_widths, hdr_color=None):
    """Styled two-tone data table."""
    hdr_color = hdr_color or C_DARK
    S = build_styles()
    data = [[Paragraph(h, S['th']) for h in headers]]
    for row in rows:
        data.append([Paragraph(str(c), S['td']) for c in row])
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0),  hdr_color),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [C_WHITE, C_BG]),
        ('GRID',          (0,0), (-1,-1), 0.5, C_BORDER),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING',   (0,0), (-1,-1), 10),
        ('RIGHTPADDING',  (0,0), (-1,-1), 10),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
    ]))
    return tbl


def draw_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(C_SUBTLE)
    canvas.drawString(20*mm, 12*mm,
        f'QA Buddy  ·  Generated {datetime.now().strftime("%d %b %Y")}')
    canvas.drawRightString(190*mm, 12*mm, f'Page {doc.page}')
    # Footer line
    canvas.setStrokeColor(C_BORDER)
    canvas.setLineWidth(0.5)
    canvas.line(20*mm, 18*mm, 190*mm, 18*mm)
    canvas.restoreState()


def make_doc(file_path):
    return SimpleDocTemplate(
        file_path, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=28*mm
    )


# ── App setup ─────────────────────────────────────────────────
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'static/uploads'

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('static/screenshots', exist_ok=True)
os.makedirs('export/screenshots', exist_ok=True)


def get_db():
    return sqlite3.connect('database.db')


# ── HOME ─────────────────────────────────────────────────────

@app.route('/')
def index():
    db = get_db()
    bug_count = db.execute("SELECT COUNT(*) FROM bugs").fetchone()[0]
    tc_count  = db.execute("SELECT COUNT(*) FROM testcases").fetchone()[0]
    open_bugs = db.execute("SELECT COUNT(*) FROM bugs WHERE status='Open'").fetchone()[0]
    passed_tc = db.execute("SELECT COUNT(*) FROM testcases WHERE status='Pass'").fetchone()[0]
    db.close()
    return render_template('index.html',
        bug_count=bug_count, tc_count=tc_count,
        open_bugs=open_bugs, passed_tc=passed_tc)


# ── BUGS ─────────────────────────────────────────────────────

@app.route('/add-bug', methods=['GET', 'POST'])
def add_bug():
    if request.method == 'POST':
        screenshot = request.files.get('screenshot')
        filename = None
        if screenshot and screenshot.filename != '':
            filename = secure_filename(screenshot.filename)
            screenshot.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        db = get_db()
        db.execute(
            "INSERT INTO bugs (title,steps,expected,actual,severity,status,screenshot,priority)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (request.form['title'], request.form['steps'], request.form['expected'],
             request.form['actual'], request.form['severity'], 'Open',
             filename, request.form['priority'])
        )
        db.commit(); db.close()
        return redirect('/bugs')
    return render_template('add_bug.html')


@app.route('/bugs')
def bugs():
    status   = request.args.get('status')
    priority = request.args.get('priority')
    query, params = "SELECT * FROM bugs WHERE 1=1", []
    if status:   query += " AND status = ?";   params.append(status)
    if priority: query += " AND priority = ?"; params.append(priority)
    db   = get_db()
    bugs = db.execute(query, params).fetchall()
    db.close()
    return render_template('bugs.html', bugs=bugs)


@app.route('/bug/<int:bug_id>')
def bug_detail(bug_id):
    db  = get_db()
    bug = db.execute("SELECT * FROM bugs WHERE id=?", (bug_id,)).fetchone()
    db.close()
    return render_template('bug_detail.html', bug=bug)


@app.route('/update-status/<int:bug_id>', methods=['POST'])
def update_status(bug_id):
    db = get_db()
    db.execute("UPDATE bugs SET status=? WHERE id=?", (request.form['status'], bug_id))
    db.commit(); db.close()
    return redirect(f'/bug/{bug_id}')


@app.route('/delete-bug/<int:bug_id>', methods=['POST'])
def delete_bug(bug_id):
    db = get_db()
    db.execute("DELETE FROM bugs WHERE id=?", (bug_id,))
    db.commit(); db.close()
    return redirect('/bugs')


@app.route('/delete-all-bugs', methods=['POST'])
def delete_all_bugs():
    db = get_db()
    db.execute("DELETE FROM bugs")
    db.commit(); db.close()
    return redirect('/bugs')


# ── TEST CASES ────────────────────────────────────────────────

@app.route('/add-testcase', methods=['GET', 'POST'])
def add_testcase():
    if request.method == 'POST':
        db = get_db()
        db.execute(
            "INSERT INTO testcases (module,scenario,steps,expected,status,technique)"
            " VALUES (?,?,?,?,?,?)",
            (request.form['module'], request.form['scenario'], request.form['steps'],
             request.form['expected'], request.form['status'], request.form['technique'])
        )
        db.commit(); db.close()
        return redirect('/testcases')
    return render_template('add_testcase.html')


@app.route('/testcases')
def testcases():
    status = request.args.get('status')
    query, params = "SELECT * FROM testcases WHERE 1=1", []
    if status: query += " AND status=?"; params.append(status)
    db        = get_db()
    testcases = db.execute(query, params).fetchall()
    db.close()
    return render_template('testcases.html', testcases=testcases)


@app.route('/testcase/<int:tc_id>')
def testcase_detail(tc_id):
    db = get_db()
    tc = db.execute("SELECT * FROM testcases WHERE id=?", (tc_id,)).fetchone()
    db.close()
    return render_template('testcase_detail.html', tc=tc)


@app.route('/update-testcase-status/<int:tc_id>', methods=['POST'])
def update_testcase_status(tc_id):
    db = get_db()
    db.execute("UPDATE testcases SET status=? WHERE id=?", (request.form['status'], tc_id))
    db.commit(); db.close()
    return redirect('/testcases')


@app.route('/delete-testcase/<int:tc_id>', methods=['POST'])
def delete_testcase(tc_id):
    db = get_db()
    db.execute("DELETE FROM testcases WHERE id=?", (tc_id,))
    db.commit(); db.close()
    return redirect('/testcases')


@app.route('/delete-all-testcases', methods=['POST'])
def delete_all_testcases():
    db = get_db()
    db.execute("DELETE FROM testcases")
    db.commit(); db.close()
    return redirect('/testcases')


# ── EXPORTS ───────────────────────────────────────────────────

@app.route('/export-bugs')
def export_bugs():
    db   = get_db()
    bugs = db.execute("SELECT * FROM bugs").fetchall()
    db.close()

    screenshots_dir = "export/screenshots"
    os.makedirs(screenshots_dir, exist_ok=True)

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Bugs Report"

    hfill  = PatternFill("solid", fgColor="18181b")
    hfont  = Font(bold=True, color="FFFFFF", size=11)
    thin   = Side(style='thin', color='E4E4E7')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["S.No","Title","Severity","Priority","Status",
               "Steps","Expected Result","Actual Result","Screenshot"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hfont; cell.fill = hfill; cell.border = bdr
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, bug in enumerate(bugs, 1):
        scr = "Not Available"
        if bug[7]:
            src = f"static/uploads/{bug[7]}"
            if os.path.exists(src):
                shutil.copy(src, f"{screenshots_dir}/{bug[7]}")
                scr = f'=HYPERLINK("screenshots/{bug[7]}","Open Screenshot")'
        ws.append([i, bug[1], bug[5], bug[8], bug[6], bug[2], bug[3], bug[4], scr])
        for cell in ws[ws.max_row]:
            cell.border = bdr
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(ws.max_row, 9).font = Font(color="0563C1", underline="single")

    for i, w in enumerate([6,28,12,12,10,35,30,30,20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    excel_path = "export/bugs_report.xlsx"
    wb.save(excel_path)

    zip_path = "bugs_export.zip"
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(excel_path, "bugs_report.xlsx")
        for f in os.listdir(screenshots_dir):
            zf.write(f"{screenshots_dir}/{f}", f"screenshots/{f}")
    return send_file(zip_path, as_attachment=True)


@app.route('/export-testcases')
def export_testcases():
    db        = get_db()
    testcases = db.execute("SELECT * FROM testcases").fetchall()
    db.close()

    wb = Workbook(); ws = wb.active; ws.title = "Test Cases"
    hfill = PatternFill("solid", fgColor="18181b")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    thin  = Side(style='thin', color='E4E4E7')
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws.append(["S.No","Module","Scenario","Test Design Technique",
                            "Steps","Expected Result","Status"]) or ws[1]:
        cell.font = hfont; cell.fill = hfill; cell.border = bdr
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, tc in enumerate(testcases, 1):
        ws.append([i, tc[1], tc[2], tc[6], tc[3], tc[4], tc[5]])
        for cell in ws[ws.max_row]:
            cell.border = bdr
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    for i, w in enumerate([6,18,30,26,40,35,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    wb.save("testcases_report.xlsx")
    return send_file("testcases_report.xlsx", as_attachment=True)


# ── DOCUMENTATION ─────────────────────────────────────────────

@app.route('/documentation')
def documentation():
    return render_template('documentation.html')

@app.route('/doc/test-plan')
def test_plan_form():
    return render_template('test_plan_form.html', data=None, edit_id=None)

@app.route('/doc/test-strategy')
def test_strategy_form():
    return render_template('test_strategy_form.html', data=None, edit_id=None)

@app.route('/doc/test-summary')
def test_summary_form():
    return render_template('test_summary_form.html', data=None, edit_id=None)


# ── TEST PLAN PDF ─────────────────────────────────────────────

@app.route('/generate-test-plan', methods=['POST'])
def generate_test_plan():
    objective   = request.form['objective']
    scope       = request.form['scope']
    environment = request.form['environment']
    risks       = request.form['risks']
    edit_id     = request.form.get('edit_id')

    data = {"objective": objective, "scope": scope,
            "environment": environment, "risks": risks}
    db = get_db()
    if edit_id:
        db.execute("UPDATE documentation SET content=? WHERE id=?",
                   (json.dumps(data), edit_id))
    else:
        db.execute("INSERT INTO documentation (doc_type,content) VALUES (?,?)",
                   ("test_plan", json.dumps(data)))
    db.commit(); db.close()

    S     = build_styles()
    story = []

    story.append(pdf_header("Test Plan",
        f"Prepared with QA Buddy  ·  {datetime.now().strftime('%d %B %Y')}",
        C_ACCENT))
    story.append(Spacer(1, 14))

    sections = [
        ("1. Objective",            objective,   C_ACCENT),
        ("2. Scope",                scope,       C_ACCENT),
        ("3. Test Environment",     environment, C_MID),
        ("4. Risks & Mitigations",  risks,       C_AMBER),
    ]
    for heading, content, accent in sections:
        story.append(Paragraph(heading, S['heading']))
        story.append(section_card(heading.split('. ',1)[1], content, accent))

    doc = make_doc("Test_Plan.pdf")
    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    return send_file("Test_Plan.pdf", as_attachment=True)


# ── TEST STRATEGY PDF ─────────────────────────────────────────

@app.route('/generate-test-strategy', methods=['POST'])
def generate_test_strategy():
    test_levels = request.form['test_levels']
    test_types  = request.form['test_types']
    techniques  = request.form['techniques']
    tools       = request.form['tools']
    metrics     = request.form['metrics']
    edit_id     = request.form.get('edit_id')

    data = {"test_levels": test_levels, "test_types": test_types,
            "techniques": techniques, "tools": tools, "metrics": metrics}
    db = get_db()
    if edit_id:
        db.execute("UPDATE documentation SET content=? WHERE id=?",
                   (json.dumps(data), edit_id))
    else:
        db.execute("INSERT INTO documentation (doc_type,content) VALUES (?,?)",
                   ("test_strategy", json.dumps(data)))
    db.commit(); db.close()

    S     = build_styles()
    story = []

    story.append(pdf_header("Test Strategy",
        f"Prepared with QA Buddy  ·  {datetime.now().strftime('%d %B %Y')}",
        C_AMBER))
    story.append(Spacer(1, 14))

    sections = [
        ("1. Test Levels",            test_levels, C_ACCENT),
        ("2. Test Types",             test_types,  C_ACCENT),
        ("3. Test Design Techniques", techniques,  C_AMBER),
        ("4. Tools Used",             tools,       C_MID),
        ("5. Quality Metrics",        metrics,     C_GREEN),
    ]
    for heading, content, accent in sections:
        story.append(Paragraph(heading, S['heading']))
        story.append(section_card(heading.split('. ',1)[1], content, accent))

    doc = make_doc("Test_Strategy.pdf")
    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    return send_file("Test_Strategy.pdf", as_attachment=True)


# ── TEST SUMMARY PDF ──────────────────────────────────────────

@app.route('/generate-test-summary', methods=['POST'])
def generate_test_summary():
    execution_summary = request.form['execution_summary']
    total_tc       = request.form['total_tc']
    passed         = request.form['passed']
    failed         = request.form['failed']
    blocked        = request.form.get('blocked', 0)
    critical       = request.form.get('critical', 0)
    high           = request.form.get('high', 0)
    medium         = request.form.get('medium', 0)
    low            = request.form.get('low', 0)
    recommendation = request.form['recommendation']
    edit_id        = request.form.get('edit_id')

    data = {
        "execution_summary": execution_summary,
        "total_tc": total_tc, "passed": passed,
        "failed": failed, "blocked": blocked,
        "critical": critical, "high": high,
        "medium": medium, "low": low,
        "recommendation": recommendation
    }
    db = get_db()
    if edit_id:
        db.execute("UPDATE documentation SET content=? WHERE id=?",
                   (json.dumps(data), edit_id))
    else:
        db.execute("INSERT INTO documentation (doc_type,content) VALUES (?,?)",
                   ("test_summary", json.dumps(data)))
    db.commit(); db.close()

    S     = build_styles()
    story = []

    story.append(pdf_header("Test Summary Report",
        f"Prepared with QA Buddy  ·  {datetime.now().strftime('%d %B %Y')}",
        C_GREEN))
    story.append(Spacer(1, 14))

    # Execution summary text
    story.append(Paragraph("1. Execution Summary", S['heading']))
    story.append(section_card("Summary", execution_summary, C_GREEN))

    # Test execution results table
    story.append(Paragraph("2. Test Execution Results", S['heading']))
    story.append(Spacer(1, 6))
    story.append(data_table(
        headers=["Metric", "Count"],
        rows=[["Total Test Cases", total_tc],
              ["Passed", passed], ["Failed", failed], ["Blocked", blocked]],
        col_widths=[120*mm, 50*mm],
        hdr_color=C_GREEN,
    ))
    story.append(Spacer(1, 14))

    # Defect severity table
    story.append(Paragraph("3. Defect Summary by Severity", S['heading']))
    story.append(Spacer(1, 6))
    story.append(data_table(
        headers=["Severity", "Count"],
        rows=[["Critical", critical], ["High", high],
              ["Medium", medium],     ["Low", low]],
        col_widths=[120*mm, 50*mm],
        hdr_color=C_RED,
    ))
    story.append(Spacer(1, 14))

    # Conclusion
    story.append(Paragraph("4. Conclusion &amp; Recommendations", S['heading']))
    story.append(section_card("Conclusion", recommendation, C_ACCENT))

    doc = make_doc("Test_Summary_Report.pdf")
    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    return send_file("Test_Summary_Report.pdf", as_attachment=True)


# ── DRAFTS ────────────────────────────────────────────────────

@app.route('/documentation/drafts')
def documentation_drafts():
    db     = get_db()
    drafts = db.execute(
        "SELECT id, doc_type, content, created_at FROM documentation ORDER BY created_at DESC"
    ).fetchall()
    db.close()
    return render_template('document_drafts.html', drafts=drafts)


@app.route('/edit-doc/<int:doc_id>')
def edit_document(doc_id):
    db  = get_db()
    row = db.execute("SELECT doc_type,content FROM documentation WHERE id=?",
                     (doc_id,)).fetchone()
    db.close()
    if not row:
        return redirect('/documentation/drafts')
    doc_type = row[0]
    data     = json.loads(row[1])
    if doc_type == "test_plan":
        return render_template('test_plan_form.html', data=data, edit_id=doc_id)
    if doc_type == "test_strategy":
        return render_template('test_strategy_form.html', data=data, edit_id=doc_id)
    if doc_type == "test_summary":
        return render_template('test_summary_form.html', data=data, edit_id=doc_id)
    return redirect('/documentation/drafts')


@app.route('/delete-all-docs', methods=['POST'])
def delete_all_docs():
    db = get_db()
    db.execute("DELETE FROM documentation")
    db.commit(); db.close()
    return redirect('/documentation/drafts')


# ── AI AGENT ─────────────────────────────────────────────────

@app.route('/ai-agent')
def ai_agent():
    return render_template('ai_agent.html')


def crawl_site(start_url, max_pages=10):
    visited, to_visit, pages = set(), [start_url], []
    domain = urlparse(start_url).netloc
    while to_visit and len(visited) < max_pages:
        url = to_visit.pop(0)
        if url in visited: continue
        try:
            res = requests.get(url, timeout=8)
            pages.append({"url": url, "status_code": res.status_code, "html": res.text})
            visited.add(url)
            for a in BeautifulSoup(res.text, "html.parser").find_all("a", href=True):
                link = urljoin(url, a["href"])
                if urlparse(link).netloc == domain and link not in visited:
                    to_visit.append(link)
        except Exception as e:
            pages.append({"url": url, "status_code": "ERROR", "error": str(e)})
            visited.add(url)
    return pages


def confidence_score(status, error_text=None):
    score = 0.6
    if status in [500, "ERROR"]: score += 0.3
    if status == 404:            score += 0.2
    if error_text:               score += 0.1
    return min(score, 1.0)


def generate_bug_summary(bug):
    return (f"The page at {bug['url']} failed due to "
            f"{bug['actual'] if bug.get('actual') else 'an HTTP error'}. "
            "This impacts user access and should be fixed.")


def capture_screenshot(url):
    os.makedirs("static/screenshots", exist_ok=True)
    filename = f"screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    path     = f"static/screenshots/{filename}"
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page    = browser.new_page()
            page.goto(url, timeout=10000, wait_until="load")
            page.screenshot(path=path, full_page=True)
            browser.close()
        return path
    except Exception as e:
        print("SCREENSHOT ERROR:", e)
        return None


def detect_bugs(pages):
    bugs = []
    for page in pages:
        url, status = page.get("url"), page.get("status_code")
        if status == "ERROR":
            b = {"title": "Page not reachable", "url": url, "severity": "Critical",
                 "steps": f"1. Open browser\n2. Navigate to {url}",
                 "expected": "Page should load successfully",
                 "actual": page.get("error"),
                 "confidence": confidence_score(status, page.get("error")),
                 "screenshot": capture_screenshot(url)}
            b["summary"] = generate_bug_summary(b); bugs.append(b)
        elif status in [403, 404, 500]:
            b = {"title": f"HTTP {status} error on page", "url": url,
                 "severity": "High" if status >= 500 else "Major",
                 "steps": f"1. Open browser\n2. Navigate to {url}",
                 "expected": "Page should load correctly",
                 "actual": f"Server returned HTTP {status}",
                 "confidence": confidence_score(status),
                 "screenshot": capture_screenshot(url)}
            b["summary"] = generate_bug_summary(b); bugs.append(b)
    return bugs


def save_ai_bugs(bugs):
    db = get_db()
    for b in bugs:
        db.execute(
            "INSERT INTO bugs (title,steps,expected,actual,severity,priority,status,screenshot)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (b["title"], b["steps"], b["expected"], b["actual"],
             b["severity"], "Medium", "Open", b.get("screenshot"))
        )
    db.commit(); db.close()


@app.route('/ai-scan', methods=['POST'])
def ai_scan():
    target_url = request.form.get('url', '').strip()
    if not target_url:
        return render_template("ai_results.html", bugs=[], scanned=0)
    pages = crawl_site(target_url)
    bugs  = detect_bugs(pages)
    if request.args.get('severity'):
        bugs = [b for b in bugs if b["severity"] == request.args['severity']]
    if request.args.get('min_conf', type=float):
        bugs = [b for b in bugs if b["confidence"] >= request.args.get('min_conf', type=float)]
    if bugs:
        save_ai_bugs(bugs)
    return render_template("ai_results.html", bugs=bugs, scanned=len(pages))


# ── ENTRY POINT ───────────────────────────────────────────────

if __name__ == '__main__':
    app.run(debug=True)